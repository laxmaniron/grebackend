import openpyxl
from openpyxl import Workbook, load_workbook
import pymongo

myclient = pymongo.MongoClient("mongodb://localhost:27017/")
mydb = myclient["GRE"]
source = mydb["grevocabs"]

# path = "./apparels_maininfo.xlsx"
path = "./GRE_words.xlsx"

wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active

inserter = {}

p = 0

count = 0

# for x in source.find():

#     # print(x["page_color_link"])

#     # print(count)

#     # commented code

#     buffer = []

for dress in range(3, (sheet_obj.max_row)+1):
    m = sheet_obj.cell(row=dress, column=1)
    print(m.value)

    inserter = {}

    word = str(m.value)

    vocab, typeofvocab = word.split("(")

    # print(vocab.strip(" "), len(vocab.strip(" ")))

    # print(word.split("("))

    # print(typeofvocab[:-2])

    inserter["wordname"] = vocab

    inserter["typeofvocab"] = typeofvocab.split(")")[0]

    meaning = sheet_obj.cell(row=dress, column=2)
    inserter["word_meaning"] = meaning.value

    mnemonic = sheet_obj.cell(row=dress, column=3)
    inserter["word_mnemonic"] = mnemonic.value

    word_example = sheet_obj.cell(row=dress, column=4)
    inserter["word_example"] = word_example.value

    inserter["origin_of_word"] = "Magoosh_GRE_High_frequency_list"
    source.insert_one(inserter)

    print(inserter)

# print("ok")
